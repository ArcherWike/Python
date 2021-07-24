from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = 5
ws["A2"] = "Hello"
ws["A3"] = "This"
ws["A4"] = "program run!"

wb.save("sample2.xlsx")