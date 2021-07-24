from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
print(wb.sheetnames)

sheet_rangers = wb['Sheet']
print(sheet_rangers['A1'].value, sheet_rangers['B2'].value, sheet_rangers['C3'].value, sheet_rangers['D4'].value)

print(sheet_rangers['A7'].value)
print(sheet_rangers['B7'].value)
print(sheet_rangers['C7'].value)